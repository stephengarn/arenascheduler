from openpyxl import Workbook, load_workbook
import pprint
import requests
import json
#download the file using the requests library

#commented out because we don't need to download a new schedule everytime during testing

#TESTING PURPOSES. TAKE OUT COMMENT DURING PROUDCTION!!!

url = 'https://cloud.rampinteractive.com/hockeyedmonton/files/Arena%20Reports/ARENA_report.xlsx'
r = requests.get(url, allow_redirects=True)
open("ARENA_reports.xlsx", "wb").write(r.content)


pp = pprint.PrettyPrinter(indent=4)

#open the excel file

wb = load_workbook(filename = "ARENA_reports.xlsx")

#find the worksheet we want (should be ARENA_report)

sheet_ranges = wb['ARENA_report']

#now we have to start building dicts
#and then we need to input into a JSON file for our 
#front facing application to understand 

# how should we organize the data? We should do the dict like this???
arenaDict = {}
for i in range(2, sheet_ranges.max_row):
    arenaDict[i] = {
        'date': sheet_ranges["A" + str(i)].value,
        'arenaCode': sheet_ranges["B" + str(i)].value,
        'startTime': sheet_ranges["C" + str(i)].value,
        'endTime': sheet_ranges["D" + str(i)].value,
        'day': sheet_ranges["E" + str(i)].value,
        'type': sheet_ranges["F"+ str(i)].value,
        'area': sheet_ranges["G" + str(i)].value,
        'division': sheet_ranges["H" + str(i)].value,
        'visitorName': sheet_ranges["I" + str(i)].value,
        'homeName': sheet_ranges["J" + str(i)].value,
        'area2HPR': sheet_ranges["K" + str(i)].value,
        'divisionHPR': sheet_ranges["L" + str(i)].value,
        'visitorNameHPR': sheet_ranges["M" + str(i)].value,
        'homeNameHPR': sheet_ranges["N" + str(i)].value
    }


with open('arenaMinorSchedule.json', 'w', encoding='utf-8') as f:
    json.dump(arenaDict, f, ensure_ascii=False, indent=4, default=str)



#pp.pprint(arenaDict)


# arenaDict = {
#    1: {
#   'date': '2021-11-05 00:00:00',
#   'arenaCode': 'BHA',
#   'startTime': '16:30:00',
#   'endTime': '18:00:00',
#   'day': 'FRI',
#   'type': 'PRA',    if type equals HPR then it is a half ice practice and we need additional columns
#   'area': 'SAC'     if arenaDict['type'] equals 'PRA' && arena['Home'] is empty string '' then arena['home'] will 
#                     be assigned to area. 
#   'division': 'U15 tier 1 NBC',
#   'Visitor name': '',   if type equals PRA then there will be no visitor property
#   'Home name':  'SW402',
#   'area2HPR': '',
#   'divisionHPR': '',
#   'visitor nameHPR': '',
#   'home nameHPR': ''
#   },
#   2: {
#
#
#   } etc.
# }

# find max row count to loop through     
# print(sheet_ranges.max_row) row count
# print(sheet_ranges.max_column) column count

# print(sheet_ranges['B1'].value) prints off the rink
# column. sheet_ranges['A1-99'] prints off the DATE
# ['C1'] is start time
# ['D1'] is end time
# ['E1'] day
# ['F1'] type
# ['G1'] area
# ['H1'] division
# ['I1'] visitor name
# ['J1'] home name
# everything pass here I think would be considered a timbit 
# hockey game and or practice.