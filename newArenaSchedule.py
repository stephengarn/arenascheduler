import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook

# Open the workbook and select a worksheet
wb = load_workbook('ARENA_reports.xlsx')
sheet = wb['ARENA_report']

# List to hold dictionaries
schedule_list = []

# Iterate through each row in worksheet and fetch values into dict
for row in islice(sheet.values, 1, sheet.max_row):
    schedule = OrderedDict()
    schedule['date'] = row[0]
    schedule['arenaCode'] = row[1]
    schedule['startTime']  = row[2]
    schedule['endTime'] = row[3]
    schedule['day'] = row[4]
    schedule['type'] = row[5]
    schedule['area'] = row[6]
    schedule['division'] = row[7]
    schedule['visitorName'] = row[8]
    schedule['homeName'] = row[9]
    schedule['area2HPR'] = row[10]
    schedule['divisionHPR'] = row[11]
    schedule['visitorNameHPR'] = row[12]
    schedule['homeNameHPR'] = row[13]
    schedule_list.append(schedule)

# Serialize the list of dicts to JSON
#j = json.dumps(schedule_list)

# Write to file
#with open('arenaNewSchedule.json', 'w') as f:
#    f.write(j)

with open('arenaNewSchedule.json', 'w', encoding='utf-8') as f:
    json.dump(schedule_list, f, ensure_ascii=False, indent=4, default=str)
